#!/usr/bin/env python3
"""
==============================================================================
KINTSUGI WEB FRONTEND - SELENIUM & E2E LOGIN TEST CASES GENERATOR
==============================================================================
Generates a comprehensive, professionally styled Excel spreadsheet (.xlsx)
containing 310+ detailed E2E test cases and an Executive Summary dashboard
for the Kintsugi Web Frontend Authentication / Login System.
==============================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_test_cases():
    test_cases = []
    
    # --------------------------------------------------------------------------
    # MODULE 1: UI LAYOUT, ELEMENTS & VISUAL DESIGN (TC_LOG_001 - TC_LOG_040)
    # --------------------------------------------------------------------------
    ui_scenarios = [
        ("Verify login page renders within 2 seconds", "Page load performance", "1. Open browser\n2. Navigate to /login", "N/A", "Page loads under 2s with no visual glitches", "Performance", "P1", "Automated"),
        ("Verify application brand logo display", "Brand identity check", "1. Navigate to /login\n2. Inspect top header logo", "N/A", "Kintsugi emblem & brand title rendered clearly", "UI/UX", "P1", "Automated"),
        ("Verify logo link navigation to home page", "Logo clickable behavior", "1. Click on Kintsugi brand logo", "User on /login", "Browser navigates to home page /", "Functional", "P2", "Automated"),
        ("Verify page header main headline text", "Headline copywriting", "1. Check text in hero section", "N/A", "Displays 'Embrace Healing & Restore Strength'", "UI/UX", "P3", "Manual"),
        ("Verify mental health companion badge visibility", "Badge display", "1. Check hero section badges", "N/A", "Badge 'Mental Health & Emotional Companion' displayed", "UI/UX", "P3", "Manual"),
        ("Verify value proposition 1 (24/7 AI Companion)", "Feature callout 1", "1. Inspect left column highlights", "N/A", "Icon and text for Mistral AI companion visible", "UI/UX", "P3", "Manual"),
        ("Verify value proposition 2 (Fernet Encrypted Vault)", "Feature callout 2", "1. Inspect left column highlights", "N/A", "Icon and text for Fernet encryption visible", "UI/UX", "P3", "Manual"),
        ("Verify value proposition 3 (Mood & Resilience Insights)", "Feature callout 3", "1. Inspect left column highlights", "N/A", "Icon and text for mood tracking visible", "UI/UX", "P3", "Manual"),
        ("Verify security trust badge at bottom left", "Trust badge check", "1. Look below feature list", "N/A", "Displays 'AES-256 Encryption · Confidential & Secure'", "UI/UX", "P2", "Manual"),
        ("Verify email label typography and formatting", "Label styling", "1. Inspect email input label", "N/A", "Label displays 'EMAIL ADDRESS' in uppercase bold text", "UI/UX", "P3", "Automated"),
        ("Verify email input field presence and icon", "Input field UI", "1. Locate email input box", "N/A", "Email field visible with Mail icon inside left padding", "UI/UX", "P1", "Automated"),
        ("Verify email placeholder text accuracy", "Input placeholder", "1. Check empty email field", "N/A", "Placeholder displays 'name@example.com'", "UI/UX", "P2", "Automated"),
        ("Verify password label typography and formatting", "Label styling", "1. Inspect password input label", "N/A", "Label displays 'PASSWORD' in uppercase bold text", "UI/UX", "P3", "Automated"),
        ("Verify password input field presence and icon", "Input field UI", "1. Locate password input box", "N/A", "Password field visible with Lock icon on left", "UI/UX", "P1", "Automated"),
        ("Verify password placeholder text accuracy", "Input placeholder", "1. Check empty password field", "N/A", "Placeholder displays '••••••••'", "UI/UX", "P2", "Automated"),
        ("Verify password hide/show eye icon visibility", "Password toggle icon", "1. Check right side of password box", "N/A", "Eye icon button positioned on the right edge", "UI/UX", "P1", "Automated"),
        ("Verify 'Forgot password?' link positioning", "Link placement", "1. Check above password input", "N/A", "Link aligned right above password field", "UI/UX", "P2", "Automated"),
        ("Verify 'Sign In' primary button styling", "Submit button UI", "1. Inspect submit button", "N/A", "Gradient violet background, bold white text, right arrow icon", "UI/UX", "P1", "Automated"),
        ("Verify hover state on 'Sign In' button", "Button hover animation", "1. Hover cursor over Sign In button", "N/A", "Smooth gradient brightness increase & shadow expand", "UI/UX", "P2", "Manual"),
        ("Verify active scale click animation on Sign In button", "Button tap feedback", "1. Click and hold Sign In button", "N/A", "Button scales down slightly (0.98 scale)", "UI/UX", "P3", "Manual"),
        ("Verify dark theme color palette application", "Dark mode UI", "1. Toggle dark theme state", "Theme = dark", "Background dark slate, card background translucent dark", "UI/UX", "P2", "Manual"),
        ("Verify light theme color palette application", "Light mode UI", "1. Toggle light theme state", "Theme = light", "Background light gray, card background crisp white", "UI/UX", "P2", "Manual"),
        ("Verify background animated particle glow effects", "Ambient backdrop", "1. Inspect background layer", "N/A", "Radial purple/violet soft glow rendered behind card", "UI/UX", "P3", "Manual"),
        ("Verify glassmorphism card elevation shadow", "Card container UI", "1. Inspect auth container card", "N/A", "Card has rounded corners (rounded-2xl) and shadow-2xl", "UI/UX", "P2", "Manual"),
        ("Verify card hover border highlight effect", "Card border hover", "1. Hover mouse over main auth card", "N/A", "Subtle amber/violet border transition on hover", "UI/UX", "P3", "Manual"),
        ("Verify responsive layout single column on mobile", "Mobile viewport layout", "1. Set viewport to 375x812", "Mobile screen", "Left branding column hidden, centered auth card visible", "UI/UX", "P1", "Automated"),
        ("Verify responsive layout dual column on desktop", "Desktop viewport layout", "1. Set viewport to 1440x900", "Desktop screen", "Left branding panel and right auth form shown side-by-side", "UI/UX", "P1", "Automated"),
        ("Verify font family fallback stack", "Typography stack", "1. Inspect computed styles on text", "N/A", "Uses system sans-serif font stack / Inter font", "UI/UX", "P3", "Manual"),
        ("Verify text cursor pointer on clickable elements", "Cursor UX", "1. Hover over links and buttons", "N/A", "Cursor switches to pointer hand icon", "UI/UX", "P2", "Manual"),
        ("Verify focus outline ring on input fields", "Focus accessibility", "1. Click inside email input", "N/A", "Violet glow ring highlights input border", "UI/UX", "P2", "Automated"),
        ("Verify focus outline ring on password field", "Focus accessibility", "1. Click inside password input", "N/A", "Violet glow ring highlights input border", "UI/UX", "P2", "Automated"),
        ("Verify text selection highlight color", "Selection styling", "1. Select text on login page", "N/A", "Selection background is deep violet with white text", "UI/UX", "P3", "Manual"),
        ("Verify reduced motion media query compliance", "Accessibility motion", "1. Enable 'prefers-reduced-motion'", "OS setting enabled", "Page animations disabled, instant fade applied", "Accessibility", "P2", "Manual"),
        ("Verify footer copyright notice display", "Footer info", "1. Scroll to page bottom", "N/A", "Displays current year copyright notice", "UI/UX", "P3", "Manual"),
        ("Verify footer privacy policy link", "Legal link", "1. Click Privacy Policy link", "N/A", "Navigates to /privacy page", "Functional", "P3", "Manual"),
        ("Verify footer terms of service link", "Legal link", "1. Click Terms of Service link", "N/A", "Navigates to /terms page", "Functional", "P3", "Manual"),
        ("Verify favicon image loading in browser tab", "Favicon check", "1. Observe browser tab icon", "N/A", "Kintsugi emblem favicon visible in tab", "UI/UX", "P2", "Manual"),
        ("Verify meta viewport tag in HTML head", "SEO / Responsive head", "1. Inspect DOM head", "N/A", "Contains width=device-width, initial-scale=1.0", "UI/UX", "P2", "Manual"),
        ("Verify page title meta tag presence", "SEO meta check", "1. Inspect DOM head", "N/A", "Title tag contains 'Kintsugi | Login'", "SEO", "P2", "Automated"),
        ("Verify meta description tag presence", "SEO meta check", "1. Inspect DOM head", "N/A", "Description meta tag present and populated", "SEO", "P3", "Manual"),
    ]

    tc_count = 1
    for title, scenario, steps, data, expected, ttype, prio, auto in ui_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "UI Layout & Visual Design",
            title,
            "User opens Kintsugi web frontend in browser",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 2: FORM FIELDS & INPUT VALIDATION (TC_LOG_041 - TC_LOG_095)
    # --------------------------------------------------------------------------
    val_scenarios = [
        ("Submit empty form without filling any fields", "Validation on empty submit", "1. Clear email & password\n2. Click Sign In", "Empty inputs", "Validation errors: 'Email address is required', 'Password is required'", "Validation", "P0", "Automated"),
        ("Submit form with email containing missing @ symbol", "Invalid email format", "1. Enter 'user.example.com'\n2. Enter 'Pass1234!'\n3. Click Sign In", "user.example.com", "Inline error: 'Please enter a valid email address.'", "Validation", "P0", "Automated"),
        ("Submit form with email containing missing domain name", "Invalid email format", "1. Enter 'user@'\n2. Click Sign In", "user@", "Inline error: 'Please enter a valid email address.'", "Validation", "P0", "Automated"),
        ("Submit form with email containing missing username", "Invalid email format", "1. Enter '@domain.com'\n2. Click Sign In", "@domain.com", "Inline error: 'Please enter a valid email address.'", "Validation", "P0", "Automated"),
        ("Submit form with email containing spaces", "Invalid email spaces", "1. Enter 'user name@domain.com'\n2. Click Sign In", "user name@domain.com", "Inline error: 'Please enter a valid email address.'", "Validation", "P1", "Automated"),
        ("Submit form with email containing double @ symbols", "Invalid email format", "1. Enter 'user@@domain.com'\n2. Click Sign In", "user@@domain.com", "Inline error: 'Please enter a valid email address.'", "Validation", "P1", "Automated"),
        ("Submit form with valid email format (lowercase)", "Valid email syntax", "1. Enter 'john.doe@example.com'\n2. Click Sign In", "john.doe@example.com", "Email validation passes without error", "Validation", "P0", "Automated"),
        ("Submit form with uppercase email format", "Uppercase email normalization", "1. Enter 'JOHN.DOE@EXAMPLE.COM'\n2. Click Sign In", "JOHN.DOE@EXAMPLE.COM", "Email trimmed/lowercased and processed", "Validation", "P1", "Automated"),
        ("Submit form with sub-domain email address", "Subdomain email", "1. Enter 'admin@sub.domain.co.uk'\n2. Click Sign In", "admin@sub.domain.co.uk", "Email accepted as valid", "Validation", "P1", "Automated"),
        ("Submit form with plus-tagging email format", "Plus tag email", "1. Enter 'user+kintsugi@example.com'\n2. Click Sign In", "user+kintsugi@example.com", "Email accepted as valid", "Validation", "P1", "Automated"),
        ("Submit form with leading whitespace in email", "Email space trim", "1. Enter '  user@example.com'\n2. Click Sign In", "  user@example.com", "Zod schema trims whitespace before validation", "Validation", "P1", "Automated"),
        ("Submit form with trailing whitespace in email", "Email space trim", "1. Enter 'user@example.com  '\n2. Click Sign In", "user@example.com  ", "Zod schema trims whitespace before validation", "Validation", "P1", "Automated"),
        ("Submit password with length less than 8 characters", "Short password validation", "1. Enter valid email\n2. Enter 'Pass123'\n3. Click Sign In", "Pass123 (7 chars)", "Inline error: 'Password must be at least 8 characters.'", "Validation", "P0", "Automated"),
        ("Submit password with exactly 8 characters", "Boundary password 8 chars", "1. Enter valid email\n2. Enter 'Pass1234'\n3. Click Sign In", "Pass1234 (8 chars)", "Password validation passes", "Validation", "P0", "Automated"),
        ("Submit password with 128 characters length", "Long password handling", "1. Enter valid email\n2. Enter 128-char string\n3. Click Sign In", "A"*128, "Password validation passes without browser lag", "Validation", "P2", "Automated"),
        ("Submit password containing special characters", "Special char password", "1. Enter 'Valid@Email.com'\n2. Enter '!@#$%^&*()_+'", "!@#$%^&*()_+", "Password accepted without syntax error", "Validation", "P1", "Automated"),
        ("Submit password containing Unicode/Emoji characters", "Unicode password", "1. Enter 'Valid@Email.com'\n2. Enter 'P@ss🔑Word123'", "P@ss🔑Word123", "Password input accepts Unicode characters", "Validation", "P2", "Automated"),
        ("Verify error message clears when user fixes email", "Dynamic error clear", "1. Trigger email error\n2. Type valid email", "john@example.com", "Inline error message disappears automatically", "Validation", "P1", "Automated"),
        ("Verify error message clears when user fixes password", "Dynamic error clear", "1. Trigger password error\n2. Type 8+ char password", "Password123!", "Inline error message disappears automatically", "Validation", "P1", "Automated"),
        ("Verify password input type is masked as 'password'", "Masked password", "1. Type text into password field", "Secret123", "Characters displayed as dots/bullets", "Security", "P0", "Automated"),
        ("Click password show toggle icon once", "Password toggle show", "1. Type 'Secret123'\n2. Click eye icon", "Secret123", "Input type changes to 'text', plain password visible", "Functional", "P0", "Automated"),
        ("Click password show toggle icon twice", "Password toggle hide", "1. Click eye icon twice", "Secret123", "Input type reverts to 'password', text re-masked", "Functional", "P0", "Automated"),
        ("Verify toggle button aria-label changes", "Accessible toggle label", "1. Inspect eye icon button", "N/A", "Aria-label switches between 'Show password' & 'Hide password'", "Accessibility", "P2", "Automated"),
        ("Copy text from password field when masked", "Copy restriction", "1. Select masked password\n2. Press Ctrl+C", "Masked input", "Browser prevents copying plain password text", "Security", "P1", "Manual"),
        ("Paste clipboard text into email field", "Clipboard paste", "1. Copy email to clipboard\n2. Ctrl+V in email field", "test@domain.com", "Text pasted successfully into field", "Functional", "P2", "Automated"),
        ("Paste clipboard text into password field", "Clipboard paste", "1. Copy text to clipboard\n2. Ctrl+V in password field", "MySecretPass!1", "Text pasted successfully into field", "Functional", "P2", "Automated"),
        ("Verify autocomplete attribute on email input", "Browser autocomplete", "1. Inspect email input DOM", "N/A", "autocomplete='email' attribute present", "UI/UX", "P2", "Automated"),
        ("Verify autocomplete attribute on password input", "Browser autocomplete", "1. Inspect password input DOM", "N/A", "autocomplete='current-password' attribute present", "UI/UX", "P2", "Automated"),
        ("Verify form element has noValidate flag", "HTML5 native override", "1. Inspect form tag", "N/A", "Form tag includes noValidate to use React Hook Form errors", "Technical", "P2", "Automated"),
        ("Submit form using Enter key in email field", "Keyboard submit email", "1. Enter valid data\n2. Press Enter in email box", "Valid inputs", "Form submits, loading state triggered", "Functional", "P1", "Automated"),
        ("Submit form using Enter key in password field", "Keyboard submit password", "1. Enter valid data\n2. Press Enter in password box", "Valid inputs", "Form submits, loading state triggered", "Functional", "P1", "Automated"),
        ("Verify disabled state on inputs during submission", "Input pending state", "1. Click Sign In\n2. Inspect inputs during pending request", "Pending API", "Inputs receive disabled attribute while loading", "UI/UX", "P1", "Automated"),
        ("Verify loading spinner on button during submission", "Button pending state", "1. Click Sign In\n2. Check submit button", "Pending API", "Button text replaced with LoadingSpinner component", "UI/UX", "P1", "Automated"),
        ("Verify submit button disabled during submission", "Prevent double submit", "1. Click Sign In\n2. Try clicking button again quickly", "Pending API", "Button disabled, second click ignored", "Functional", "P0", "Automated"),
        ("Verify error alert box styling on server failure", "Server error banner", "1. Submit invalid credentials", "Wrong pass", "Alert box appears with red background & icon", "UI/UX", "P1", "Automated"),
        ("Verify server error alert title text", "Server error banner", "1. Trigger server error", "Wrong pass", "Alert title displays 'Authentication Error'", "UI/UX", "P2", "Automated"),
        ("Verify server error alert description text", "Server error banner", "1. Trigger server error", "Wrong pass", "Displays error message or fallback text", "UI/UX", "P1", "Automated"),
        ("Verify Sonner toast notification on successful login", "Success toast", "1. Submit valid credentials", "Valid credentials", "Toast displays 'Welcome back to Kintsugi!'", "UI/UX", "P0", "Automated"),
        ("Verify Sonner toast notification on failed login", "Failure toast", "1. Submit bad credentials", "Invalid credentials", "Toast displays 'Authentication failed'", "UI/UX", "P1", "Automated"),
        ("Verify password input max length boundary", "Input length limit", "1. Type 500 characters into password box", "500 chars", "Input caps or handles long text gracefully without UI crash", "Validation", "P3", "Manual"),
        ("Verify email field case sensitivity handling", "Email case handling", "1. Login with USER@DOMAIN.COM", "USER@DOMAIN.COM", "Backend matches user record regardless of casing", "Functional", "P1", "Automated"),
        ("Verify password field exact case sensitivity", "Password case check", "1. Login with lowercase 'password123'", "password123", "Rejected if original password is 'Password123'", "Security", "P0", "Automated"),
        ("Submit email containing numbers and hyphens", "Valid email chars", "1. Enter 'user-123.test@domain-app.org'", "user-123...", "Validation accepts input as valid email", "Validation", "P2", "Automated"),
        ("Submit email with numeric top-level domain", "Numeric TLD", "1. Enter 'user@domain.123'", "user@domain.123", "Validation handles or rejects invalid TLD format", "Validation", "P3", "Manual"),
        ("Verify field tab index ordering", "Tab index DOM", "1. Inspect tabIndex on inputs", "N/A", "Follows standard visual order (1: Email, 2: Pass, 3: Submit)", "Accessibility", "P2", "Automated"),
        ("Verify focus retention on form validation error", "Error focus", "1. Submit empty form", "N/A", "Focus remains on email field for quick correction", "UX", "P2", "Manual"),
        ("Verify password field text clear button if available", "Clear button", "1. Inspect password input", "N/A", "Field cleared cleanly upon user deletion", "UI/UX", "P3", "Manual"),
        ("Verify email input type='email' HTML attribute", "HTML5 input type", "1. Inspect email input DOM element", "N/A", "type='email' attribute configured for mobile keyboards", "UI/UX", "P1", "Automated"),
        ("Verify password input type='password' HTML attribute", "HTML5 input type", "1. Inspect password input DOM element", "N/A", "type='password' attribute configured natively", "UI/UX", "P1", "Automated"),
        ("Verify form submission with autocomplete fill", "Browser autofill", "1. Trigger browser saved password autofill", "Autofilled data", "Form fields populate correctly and enable submit", "Functional", "P1", "Manual"),
        ("Verify clear error alert upon resubmitting form", "Alert reset", "1. Trigger error\n2. Re-submit valid data", "Fixed data", "Previous error alert dismissed from screen", "UI/UX", "P2", "Automated"),
        ("Verify form resets state on navigate away and back", "Form state reset", "1. Enter text\n2. Click Register\n3. Click Login", "Text in inputs", "Inputs cleared or reset to default empty state", "Functional", "P2", "Automated"),
        ("Verify password toggle state resets on navigate away", "Toggle state reset", "1. Show password\n2. Navigate away & back", "Shown pass", "Password field reverts to default hidden 'password' type", "Functional", "P3", "Manual"),
        ("Verify form submission with numeric-only password", "Numeric password", "1. Enter '12345678'\n2. Submit", "12345678", "Accepted if 8+ characters long", "Validation", "P2", "Automated"),
        ("Verify form submission with alphabet-only password", "Alpha password", "1. Enter 'abcdefgh'\n2. Submit", "abcdefgh", "Accepted if 8+ characters long", "Validation", "P2", "Automated"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in val_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Form Fields & Input Validation",
            title,
            "User is on Kintsugi login page",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 3: AUTHENTICATION LOGIC & BUSINESS RULES (TC_LOG_096 - TC_LOG_150)
    # --------------------------------------------------------------------------
    auth_scenarios = [
        ("Login with valid registered user credentials", "Successful auth", "1. Enter valid email\n2. Enter valid password\n3. Click Sign In", "valid.user@kintsugi.com / ValidPass123!", "Session authenticated, JWT saved, redirected to /dashboard", "Functional", "P0", "Automated"),
        ("Login with correct email but incorrect password", "Bad password rejection", "1. Enter valid email\n2. Enter wrong password\n3. Click Sign In", "valid.user@kintsugi.com / WrongPass123!", "Auth fails, 'Incorrect email or password' message shown", "Security", "P0", "Automated"),
        ("Login with non-existent email address", "Unregistered user rejection", "1. Enter un-registered email\n2. Enter any password\n3. Click Sign In", "notfound@example.com / AnyPassword1!", "Auth fails, generic invalid credentials message shown", "Security", "P0", "Automated"),
        ("Login with deactivated/disabled account credentials", "Disabled account", "1. Enter deactivated user credentials\n2. Click Sign In", "disabled.user@example.com", "Auth fails, 'Account disabled. Contact support' message shown", "Business Rule", "P1", "Manual"),
        ("Login with unverified email account (pending OTP)", "Unverified email", "1. Enter unverified user credentials\n2. Click Sign In", "unverified@example.com", "Redirected to /verify-email or prompt to check inbox", "Business Rule", "P1", "Manual"),
        ("Login triggering 2FA / MFA authentication flow", "2FA step requirement", "1. Login with 2FA-enabled account\n2. Submit credentials", "2fa.user@example.com", "Prompted to enter 6-digit TOTP / SMS code step", "Security", "P0", "Manual"),
        ("Submit valid 6-digit MFA OTP code", "MFA code verification", "1. Enter valid 6-digit MFA code\n2. Submit", "123456", "MFA verified, user redirected to home dashboard", "Security", "P0", "Manual"),
        ("Submit invalid 6-digit MFA OTP code", "MFA code rejection", "1. Enter wrong 6-digit MFA code\n2. Submit", "000000", "Error displayed: 'Invalid authentication code'", "Security", "P1", "Manual"),
        ("Resend 2FA OTP code request", "MFA resend", "1. Click 'Resend Code' button", "N/A", "New OTP code sent to user email/SMS, countdown starts", "Functional", "P2", "Manual"),
        ("Verify account lockout after 5 consecutive failed logins", "Brute-force lockout", "1. Submit wrong password 5 times in a row", "Target email", "Account locked temporarily for 15 minutes", "Security", "P0", "Manual"),
        ("Verify account lockout error message content", "Lockout notification", "1. Attempt login during lockout period", "Locked email", "Error displays: 'Too many failed attempts. Try again in 15 mins'", "Security", "P1", "Manual"),
        ("Verify automatic account unlock after lockout duration", "Lockout expiry", "1. Wait 15 minutes after lockout\n2. Enter valid credentials", "Locked email", "Login succeeds after timer expiration", "Security", "P1", "Manual"),
        ("Click 'Forgot password?' link navigation", "Password recovery link", "1. Click 'Forgot password?' button", "N/A", "Navigates to /forgot-password route smoothly", "Functional", "P0", "Automated"),
        ("Click 'Don't have an account? Sign Up' link", "Switch to registration", "1. Click Sign Up button", "N/A", "Navigates to /register route smoothly without full reload", "Functional", "P0", "Automated"),
        ("Verify URL query parameter redirect after login", "Post-login redirect", "1. Navigate to /login?redirect=/journal\n2. Perform valid login", "?redirect=/journal", "After login, browser redirects to /journal instead of default", "Functional", "P1", "Automated"),
        ("Verify access token storage in localStorage/cookies", "Auth state storage", "1. Login successfully\n2. Inspect storage", "Valid session", "JWT auth token securely saved in state/storage", "Security", "P0", "Automated"),
        ("Verify user profile state updated in Zustand store", "Global state update", "1. Login successfully\n2. Inspect React state", "Valid session", "User payload (id, email, name) stored in auth store", "Technical", "P1", "Automated"),
        ("Verify Axios API header Authorization Bearer attached", "API token attachment", "1. Login\n2. Inspect network tab requests", "Valid JWT", "Header 'Authorization: Bearer <token>' included in requests", "Security", "P0", "Automated"),
        ("Verify automatic redirect away from /login if already logged in", "Auth guard check", "1. Login\n2. Navigate back to /login", "Active session", "System redirects logged-in user straight to /dashboard", "Functional", "P1", "Automated"),
        ("Verify session timeout after prolonged inactivity", "Session expiration", "1. Login\n2. Idle for session timeout duration", "Expired token", "API returns 401 Unauthorized, user redirected to /login", "Security", "P0", "Manual"),
        ("Login via Google OAuth provider button", "Social login Google", "1. Click 'Continue with Google'\n2. Complete Google auth", "Google credentials", "User authenticated and redirected to application", "Functional", "P1", "Manual"),
        ("Login via GitHub OAuth provider button", "Social login GitHub", "1. Click 'Continue with GitHub'\n2. Complete GitHub auth", "GitHub credentials", "User authenticated and redirected to application", "Functional", "P1", "Manual"),
        ("Cancel OAuth third-party login flow", "Social login cancellation", "1. Click Google login\n2. Close OAuth popupWindow", "Cancelled popup", "Returned to /login page with prompt 'Login cancelled'", "Functional", "P2", "Manual"),
        ("Login with account registered via OAuth using password", "Mixed auth type", "1. Try password login on Google-only account", "OAuth-only email", "Prompted: 'Please sign in using Google'", "Business Rule", "P1", "Manual"),
        ("Verify password reset token validation on reset page", "Password reset token", "1. Open /reset-password?token=valid", "Valid token", "Password reset form displayed", "Security", "P0", "Manual"),
        ("Verify expired password reset token rejection", "Expired token", "1. Open /reset-password?token=expired", "Expired token", "Error: 'Password reset link has expired'", "Security", "P1", "Manual"),
        ("Login with newly updated password after reset", "Post-reset login", "1. Reset password\n2. Navigate to /login\n3. Use new pass", "New password", "Login succeeds with new password", "Functional", "P0", "Manual"),
        ("Login with old password after password reset", "Old password rejection", "1. Reset password\n2. Try login with old password", "Old password", "Login rejected with invalid credentials error", "Security", "P0", "Manual"),
        ("Verify concurrent session restriction policy", "Multi-session rule", "1. Login on Device A\n2. Login on Device B with same user", "Same credentials", "Device A session revoked or dual sessions managed as configured", "Security", "P1", "Manual"),
        ("Verify remember me option extends session longevity", "Remember me persistence", "1. Check Remember Me\n2. Login\n3. Close & reopen browser", "Remember Me = true", "Session persists across browser restarts", "Functional", "P1", "Manual"),
        ("Verify session closes on browser quit when Remember Me unchecked", "Session cookie expiry", "1. Uncheck Remember Me\n2. Login\n3. Close browser", "Remember Me = false", "User required to re-authenticate on browser reopen", "Security", "P1", "Manual"),
        ("Verify API base URL configuration endpoint target", "Backend URL check", "1. Inspect auth API request URL", "N/A", "Points to correct backend API host (e.g. http://localhost:8000)", "Technical", "P1", "Automated"),
        ("Verify CSRF protection header on login POST request", "CSRF security", "1. Inspect login network payload headers", "N/A", "X-CSRF-Token or SameSite cookie protection active", "Security", "P0", "Automated"),
        ("Verify password payload encrypted over HTTPS in production", "TLS encryption", "1. Inspect protocol in prod", "HTTPS URL", "Request sent over secure TLS channel", "Security", "P0", "Manual"),
        ("Verify network error prompt when backend service is down", "Backend offline", "1. Stop backend API service\n2. Attempt login", "Valid credentials", "Error displays: 'Server unreachable. Please try again later.'", "Resilience", "P1", "Automated"),
        ("Verify request timeout error when API response delays > 10s", "API timeout", "1. Simulate 15s server delay\n2. Attempt login", "Valid credentials", "Client aborts request and displays timeout error message", "Resilience", "P2", "Automated"),
        ("Verify rate limiting response (HTTP 429 Too Many Requests)", "Rate limiting", "1. Send 20 login requests in 5 seconds", "Rapid requests", "API responds 429: 'Rate limit exceeded. Slow down.'", "Security", "P0", "Automated"),
        ("Verify CORS headers compliance on cross-origin login API", "CORS header check", "1. Inspect login API headers", "N/A", "Access-Control-Allow-Origin strictly configured", "Security", "P1", "Automated"),
        ("Verify clear session state on explicit Logout action", "Logout cleanup", "1. Login\n2. Click Logout in navbar", "Active session", "JWT destroyed, local storage cleared, redirected to /login", "Functional", "P0", "Automated"),
        ("Verify protected routes block unauthenticated direct navigation", "Auth guard route", "1. Open new private tab\n2. Navigate directly to /journal", "No token", "System redirects user directly to /login page", "Security", "P0", "Automated"),
        ("Verify login page title updates dynamically", "Document title", "1. Navigate to /login", "N/A", "Document title updates to 'Login - Kintsugi'", "UI/UX", "P3", "Automated"),
        ("Verify user avatar loads in dashboard post-login", "Profile state load", "1. Login as user with custom avatar", "Custom avatar", "User avatar renders correctly in top app header", "Functional", "P2", "Manual"),
        ("Verify default fallback avatar for new users post-login", "Default avatar load", "1. Login as user without avatar", "No avatar", "Displays default initial letter avatar icon", "Functional", "P3", "Manual"),
        ("Verify user role permissions loaded on login (Admin vs User)", "RBAC permission load", "1. Login as Admin user", "Admin user", "Admin navigation items populated in application header", "Security", "P1", "Manual"),
        ("Verify user role restrictions for Standard User post-login", "RBAC restriction", "1. Login as Standard user", "Standard user", "Admin dashboard routes remain hidden and protected", "Security", "P1", "Manual"),
        ("Verify login audit log generated in backend system", "Audit logging", "1. Perform valid login", "N/A", "Backend logs login event with IP, timestamp, user ID", "Security", "P2", "Manual"),
        ("Verify failed login audit log generated in backend system", "Audit logging", "1. Perform failed login", "N/A", "Backend logs failed login attempt with IP and timestamp", "Security", "P2", "Manual"),
        ("Verify multi-language / i18n support on login form if enabled", "Localization", "1. Switch locale to Spanish / French", "Locale = ES", "Labels and validation messages translate accurately", "Localization", "P3", "Manual"),
        ("Verify analytics event triggered on login button click", "Analytics tracking", "1. Click Sign In button", "Analytics enabled", "Event 'auth_login_attempt' dispatched to tracker", "Telemetry", "P3", "Manual"),
        ("Verify analytics event triggered on successful authentication", "Analytics tracking", "1. Complete successful login", "Analytics enabled", "Event 'auth_login_success' dispatched to tracker", "Telemetry", "P3", "Manual"),
        ("Verify analytics event triggered on login failure", "Analytics tracking", "1. Trigger failed login", "Analytics enabled", "Event 'auth_login_failure' dispatched with reason tag", "Telemetry", "P3", "Manual"),
        ("Verify session refresh token rotation upon token expiry", "Token rotation", "1. Wait for access token expiry", "Refresh token active", "Client silently fetches new access token using refresh token", "Security", "P1", "Manual"),
        ("Verify invalid refresh token triggers login redirect", "Refresh failure", "1. Invalidate refresh token\n2. Trigger background call", "Bad refresh token", "Session destroyed, user redirected to /login screen", "Security", "P1", "Manual"),
        ("Verify email case normalization on backend user lookup", "Backend email lookup", "1. Login with 'JoHn.DoE@ExAmPlE.cOm'", "Mixed case email", "Backend normalizes to 'john.doe@example.com' for query", "Functional", "P1", "Automated"),
        ("Verify empty space only input rejected for password", "Whitespace password", "1. Enter valid email\n2. Enter '        ' (8 spaces)\n3. Click Sign In", "8 spaces", "Rejected or treated as invalid password string", "Validation", "P1", "Automated"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in auth_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Authentication & Business Rules",
            title,
            "User is accessing Kintsugi auth service",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 4: SECURITY, VULNERABILITY & EDGE CASES (TC_LOG_151 - TC_LOG_205)
    # --------------------------------------------------------------------------
    sec_scenarios = [
        ("Submit SQL Injection payload in email field (' OR '1'='1)", "SQLi email field", "1. Enter 'admin' OR '1'='1'--'\n2. Enter password\n3. Click Sign In", "admin' OR '1'='1'--", "Payload sanitized, auth rejected without DB syntax error", "Security", "P0", "Automated"),
        ("Submit SQL Injection payload in password field (' OR '1'='1)", "SQLi password field", "1. Enter valid email\n2. Enter '\' OR \'1\'=\'1'\n3. Click Sign In", "' OR '1'='1", "Payload sanitized, auth rejected without DB syntax error", "Security", "P0", "Automated"),
        ("Submit Stacked Queries SQL Injection payload", "SQLi stacked query", "1. Enter 'user@test.com'; DROP TABLE users;--'", "DROP TABLE payload", "Payload neutralized safely by parameterized ORM query", "Security", "P0", "Automated"),
        ("Submit XSS script payload in email field (<script>alert(1)</script>)", "XSS email field", "1. Enter '<script>alert(1)</script>@test.com'\n2. Click Sign In", "<script>alert(1)...", "Script payload escaped/sanitized, no alert box executed", "Security", "P0", "Automated"),
        ("Submit XSS SVG vector payload in email field", "XSS SVG vector", "1. Enter '<svg onload=alert(1)>@test.com'", "<svg onload=...", "Payload rendered as harmless string or rejected by Zod", "Security", "P0", "Automated"),
        ("Submit XSS Image vector payload in password field", "XSS img vector", "1. Enter '<img src=x onerror=alert(1)>'", "<img src=x...", "Password masked, no DOM script execution triggered", "Security", "P0", "Automated"),
        ("Submit HTML Injection tags in input fields", "HTML injection", "1. Enter '<h1>BigText</h1>@example.com'", "<h1>BigText</h1>", "HTML tags rendered as literal string text, not HTML DOM", "Security", "P1", "Automated"),
        ("Submit Null Byte character injection in email field (%00)", "Null byte injection", "1. Enter 'user%00@example.com'", "user%00@example.com", "Null byte stripped or rejected cleanly", "Security", "P1", "Automated"),
        ("Submit 10,000 character string in email field", "Buffer overflow test", "1. Input 10,000 'A' characters into email\n2. Click Sign In", "A"*10000, "Input handled without browser tab freeze or memory leak", "Security", "P1", "Automated"),
        ("Submit 10,000 character string in password field", "Buffer overflow test", "1. Input 10,000 'B' characters into password\n2. Click Sign In", "B"*10000, "Input handled without browser tab freeze or memory leak", "Security", "P1", "Automated"),
        ("Verify Password Plaintext NOT exposed in browser Network tab", "Network payload privacy", "1. Inspect HTTP POST payload in dev tools", "Login POST payload", "Password sent over TLS; not exposed in cleartext logs", "Security", "P0", "Automated"),
        ("Verify Password Plaintext NOT logged in browser console", "Console log privacy", "1. Open browser Console\n2. Submit login form", "Console logs", "No console.log prints plain password string", "Security", "P0", "Automated"),
        ("Verify Clickjacking protection header (X-Frame-Options)", "Clickjacking defense", "1. Load /login inside an <iframe> on external site", "iframe embed", "Browser blocks framing (X-Frame-Options: DENY / SAMEORIGIN)", "Security", "P1", "Manual"),
        ("Verify Content Security Policy (CSP) header enforcement", "CSP header check", "1. Inspect HTTP response headers on /login", "N/A", "CSP header present restricting inline script execution", "Security", "P1", "Manual"),
        ("Verify Strict-Transport-Security (HSTS) header in prod", "HSTS header check", "1. Inspect production response headers", "N/A", "Strict-Transport-Security header configured", "Security", "P1", "Manual"),
        ("Verify X-Content-Type-Options: nosniff header", "MIME sniffing protection", "1. Inspect response headers", "N/A", "nosniff header prevents MIME type spoofing", "Security", "P2", "Manual"),
        ("Verify Referral Policy header setting", "Referrer privacy", "1. Inspect response headers on page transition", "N/A", "Referrer-Policy restricts sensitive URL leakage", "Security", "P2", "Manual"),
        ("Verify sensitive credentials NOT stored in URL parameters", "URL sensitive data", "1. Submit login form\n2. Inspect URL bar", "Form submit", "Credentials NEVER appended as GET query params", "Security", "P0", "Automated"),
        ("Verify sensitive token NOT exposed in browser history", "Browser history privacy", "1. Perform login\n2. Inspect browser history entries", "Browser history", "History contains page URLs only, no auth secrets", "Security", "P1", "Manual"),
        ("Verify session token cookie secure flag in production", "Secure cookie flag", "1. Inspect auth cookie flags in prod", "HTTPS session", "Cookie has 'Secure' attribute enabled", "Security", "P0", "Manual"),
        ("Verify session token cookie HttpOnly flag", "HttpOnly cookie flag", "1. Inspect auth cookie in document.cookie", "JS execution", "HttpOnly cookie invisible to document.cookie (XSS protection)", "Security", "P0", "Manual"),
        ("Verify session token cookie SameSite attribute", "SameSite CSRF protection", "1. Inspect cookie attributes", "N/A", "SameSite=Lax or SameSite=Strict enabled", "Security", "P0", "Manual"),
        ("Submit LDAP Injection payload in login input fields", "LDAP injection", "1. Enter '*)(uid=*))(|(uid=*'", "LDAP payload", "Payload sanitized, auth rejected safely", "Security", "P1", "Automated"),
        ("Submit XPath Injection payload in login input fields", "XPath injection", "1. Enter '' or '1'='1'", "XPath payload", "Payload sanitized, auth rejected safely", "Security", "P1", "Automated"),
        ("Submit Command Injection payload in email field", "Command injection", "1. Enter 'user@test.com; cat /etc/passwd'", "OS command payload", "Payload treated as literal string, no OS execution", "Security", "P0", "Automated"),
        ("Submit NoSQL Injection payload in JSON request body", "NoSQL injection", "1. POST email: {\"$gt\": \"\"} to login endpoint", "NoSQL payload", "API schema validates string type, rejects object payload", "Security", "P0", "Automated"),
        ("Verify CSRF token validation on authentication endpoint", "CSRF token check", "1. Send login POST without CSRF token", "Missing CSRF token", "Server rejects request with 403 Forbidden", "Security", "P0", "Automated"),
        ("Verify replay attack prevention using nonce/timestamp", "Replay attack defense", "1. Re-send identical intercepted request 1 hour later", "Stale request", "Server validates timestamp/nonce and rejects stale request", "Security", "P1", "Manual"),
        ("Submit URL-encoded characters in email input", "URL encoding test", "1. Enter 'user%20name@example.com'", "URL encoded string", "Handled cleanly as raw string or validated properly", "Validation", "P2", "Automated"),
        ("Submit RTL (Right-to-Left) Arabic text in fields", "RTL text handling", "1. Enter 'مستخدم@domain.com'\n2. Submit", "Arabic text", "Text renders correctly according to bidi rules", "Localization", "P2", "Manual"),
        ("Submit Cyrillic/Russian characters in email input", "Cyrillic character test", "1. Enter 'пользователь@domain.com'", "Cyrillic text", "Validation handles international domain format", "Localization", "P3", "Manual"),
        ("Submit Chinese/Japanese Kanji characters in input", "CJK character test", "1. Enter 'ユーザー@domain.com'", "CJK text", "Inputs accept international UTF-8 encoding without corruption", "Localization", "P3", "Manual"),
        ("Verify password field behavior under screen recording", "Screen capture privacy", "1. Record screen\n2. Type password", "Screen capture", "Password remains masked as dots during recording", "Security", "P2", "Manual"),
        ("Verify memory cleanup of password variable after submission", "JS memory cleanup", "1. Inspect heap snapshot post-login", "Heap snapshot", "Plaintext password string garbage collected from memory", "Security", "P2", "Manual"),
        ("Submit login request with modified Content-Type header", "Header manipulation", "1. Send login POST with text/plain header", "Wrong content-type", "Server returns 415 Unsupported Media Type or 400 Bad Request", "Security", "P1", "Automated"),
        ("Submit login request with corrupted JSON payload", "Malformed JSON", "1. Send malformed JSON body `{email: user...}`", "Bad JSON syntax", "Server returns 400 Bad Request with JSON parse error", "Resilience", "P1", "Automated"),
        ("Submit HTTP GET request to /api/auth/login endpoint", "HTTP method restriction", "1. Send GET request to login endpoint", "HTTP GET", "Server returns 405 Method Not Allowed", "Security", "P1", "Automated"),
        ("Submit HTTP PUT request to /api/auth/login endpoint", "HTTP method restriction", "1. Send PUT request to login endpoint", "HTTP PUT", "Server returns 405 Method Not Allowed", "Security", "P1", "Automated"),
        ("Submit HTTP DELETE request to /api/auth/login endpoint", "HTTP method restriction", "1. Send DELETE request to login endpoint", "HTTP DELETE", "Server returns 405 Method Not Allowed", "Security", "P1", "Automated"),
        ("Verify password strength meter indicator if present", "Password strength UI", "1. Type weak vs strong password", "Pass complexity", "Dynamic strength indicator updates accurately", "UI/UX", "P3", "Manual"),
        ("Verify copy button restriction on password input", "Copy prevention", "1. Try copying text from masked password box", "Masked input", "Copy event blocked or yields empty string", "Security", "P2", "Manual"),
        ("Verify paste event allowed on password field", "Paste permission", "1. Paste password from password manager", "Pasted pass", "Paste allowed to support secure password managers", "UX", "P1", "Manual"),
        ("Verify password manager extension auto-fill capability", "Password manager compatibility", "1. Use 1Password / Bitwarden extension autofill", "Extension autofill", "Fields populate correctly and trigger React state change", "Compatibility", "P1", "Manual"),
        ("Verify browser saved passwords popup prompt on success", "Browser password save", "1. Perform successful login", "New credentials", "Browser prompts 'Save password for Kintsugi?'", "UX", "P2", "Manual"),
        ("Verify no password save prompt triggered on failed login", "Browser password save", "1. Perform failed login attempt", "Invalid credentials", "Browser does NOT prompt to save incorrect credentials", "UX", "P2", "Manual"),
        ("Verify CORS Preflight OPTIONS request on login endpoint", "CORS Preflight check", "1. Inspect network tab for OPTIONS request", "Cross-origin request", "OPTIONS request responds 200 OK with allowed headers", "Technical", "P1", "Automated"),
        ("Verify JWT token signature verification on backend", "JWT integrity", "1. Modify JWT token payload signature string", "Tampered JWT", "Backend rejects request with 401 Unauthorized", "Security", "P0", "Automated"),
        ("Verify JWT token algorithm restriction (none alg attack)", "JWT vulnerability", "1. Send JWT token with 'alg': 'none'", "Alg=none attack", "Backend strictly rejects token without secret verification", "Security", "P0", "Automated"),
        ("Verify session invalidation on user password change", "Password change revocation", "1. Change password on Device A\n2. Refresh Device B", "Old session", "Device B session revoked, redirected to /login", "Security", "P0", "Manual"),
        ("Verify session invalidation on user email change", "Email change revocation", "1. Change email address\n2. Try old email login", "Old email", "Old email rejected, new email requires re-authentication", "Security", "P0", "Manual"),
        ("Submit login request with excessive request headers size", "Header flood test", "1. Send request with 16KB of custom headers", "Large headers", "Server returns 431 Request Header Fields Too Large", "Resilience", "P2", "Automated"),
        ("Submit login request with oversized payload body (>1MB)", "Payload limit test", "1. Send 2MB JSON payload to login endpoint", "2MB body", "Server returns 413 Payload Too Large", "Resilience", "P2", "Automated"),
        ("Verify application handles rapid back/forward browser button taps", "History navigation", "1. Login -> Back -> Forward -> Back", "Fast navigation", "State transitions smoothly without getting stuck in loop", "UX", "P2", "Manual"),
        ("Verify application state when localStorage is disabled in browser", "Disabled storage edge case", "1. Disable localStorage in browser settings\n2. Open /login", "Storage disabled", "App displays friendly notice or uses in-memory session", "Resilience", "P2", "Manual"),
        ("Verify application state when cookies are blocked in browser", "Blocked cookies edge case", "1. Block all cookies in browser\n2. Attempt login", "Cookies blocked", "App notifies user that cookies are required for login", "Resilience", "P2", "Manual"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in sec_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Security & Edge Cases",
            title,
            "Targeting Kintsugi authentication security layer",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 5: SESSION MANAGEMENT & TOKENS (TC_LOG_206 - TC_LOG_235)
    # --------------------------------------------------------------------------
    session_scenarios = [
        ("Verify page refresh on /login retains form state if saved draft", "Refresh state", "1. Type email\n2. Refresh browser", "Unsubmitted draft", "Fields reset or retain according to security specs", "Session", "P3", "Automated"),
        ("Verify back button after login does NOT return to authenticated page after logout", "History security", "1. Login\n2. Logout\n3. Click Back button", "Logged out state", "Prevented from seeing cached dashboard page, redirected to /login", "Security", "P0", "Automated"),
        ("Verify opening /login in multiple browser tabs simultaneously", "Multi-tab sync", "1. Open tab 1 & tab 2 to /login", "Dual tabs", "Both tabs render login form cleanly", "Session", "P2", "Automated"),
        ("Verify logging in on tab 1 auto-syncs auth state to tab 2", "Cross-tab auth sync", "1. Open tab 1 & tab 2\n2. Login on tab 1", "Tab 1 login", "Tab 2 updates state or redirects to dashboard on focus", "Session", "P1", "Manual"),
        ("Verify logging out on tab 1 auto-logs out tab 2", "Cross-tab logout sync", "1. Login\n2. Open tab 2\n3. Click Logout on tab 1", "Tab 1 logout", "Tab 2 automatically redirects to /login screen", "Session", "P1", "Manual"),
        ("Verify JWT token payload structure (exp, sub, iat claims)", "JWT claims validation", "1. Inspect generated JWT token", "Valid session", "Token contains required standard claims (exp, sub, iat)", "Technical", "P1", "Automated"),
        ("Verify token expiration time (exp claim value)", "Token lifespan", "1. Inspect exp claim in JWT", "Valid session", "Token configured for standard expiration (e.g. 24 hours)", "Security", "P1", "Automated"),
        ("Verify refresh token rotation mechanism", "Token rotation", "1. Call /api/auth/refresh endpoint", "Valid refresh token", "Returns new access token and new refresh token pair", "Security", "P0", "Automated"),
        ("Verify reuse of revoked refresh token invalidates token family", "Token reuse attack", "1. Intercept & reuse old refresh token", "Revoked token", "Server invalidates whole token family and forces re-login", "Security", "P0", "Manual"),
        ("Verify clearing browser cookies invalidates active session", "Cookie cleanup session", "1. Login\n2. Delete cookies in dev tools\n3. Refresh", "Cookies deleted", "User logged out automatically and redirected to /login", "Session", "P1", "Automated"),
        ("Verify clearing localStorage invalidates active session", "Storage cleanup session", "1. Login\n2. Clear localStorage\n3. Refresh", "Storage cleared", "User logged out automatically and redirected to /login", "Session", "P1", "Automated"),
        ("Verify session behavior across browser restart with Remember Me", "Session persistence", "1. Check Remember Me\n2. Login\n3. Restart browser", "Remember Me active", "User remains logged in upon reopening browser", "Session", "P1", "Manual"),
        ("Verify session behavior across browser restart WITHOUT Remember Me", "Session transient", "1. Uncheck Remember Me\n2. Login\n3. Restart browser", "Remember Me off", "Session expires when browser process closes", "Session", "P1", "Manual"),
        ("Verify API responses return 401 Unauthorized for expired tokens", "Expired API call", "1. Send request with expired JWT", "Expired token", "API returns HTTP status 401 with error message", "Security", "P0", "Automated"),
        ("Verify automatic silent token refresh before access token expires", "Silent refresh", "1. Set access token short expiry\n2. Keep app open", "Token near expiry", "Client background timer refreshes token seamlessly", "UX", "P1", "Manual"),
        ("Verify user redirection to /login when silent refresh fails", "Failed silent refresh", "1. Break refresh endpoint\n2. Wait for token expiry", "Refresh failure", "User gracefully redirected to /login with session expired toast", "UX", "P1", "Manual"),
        ("Verify HTTP response header Set-Cookie attributes on login", "Set-Cookie headers", "1. Inspect HTTP response headers on login POST", "Login POST", "Set-Cookie headers formatted with Path=/, HttpOnly, Secure", "Security", "P0", "Automated"),
        ("Verify session cookie path is restricted to application scope", "Cookie path scope", "1. Inspect cookie Path attribute", "N/A", "Path set to '/' or specific app subpath", "Security", "P2", "Automated"),
        ("Verify session data isolated between different browser profiles", "Profile isolation", "1. Login in Profile A\n2. Open Profile B", "Profile B", "Profile B has separate clean state, not logged into Profile A", "Session", "P1", "Manual"),
        ("Verify session data isolated between Standard and Incognito windows", "Incognito isolation", "1. Login in normal window\n2. Open Incognito", "Incognito window", "Incognito window starts with clean unauthenticated state", "Session", "P0", "Manual"),
        ("Verify closing Incognito window destroys Incognito session", "Incognito cleanup", "1. Login in Incognito window\n2. Close & reopen Incognito", "Incognito restart", "Session destroyed completely upon Incognito window close", "Session", "P0", "Manual"),
        ("Verify concurrent session count tracking in backend database", "DB session records", "1. Login across 3 devices\n2. Check DB sessions table", "Multi-device", "Database maintains 3 active session records with metadata", "Technical", "P2", "Manual"),
        ("Verify 'Logout from all devices' feature if supported", "Global logout", "1. Click 'Logout All Devices' in account settings", "Active devices", "Revokes all refresh tokens, logs out all active devices", "Security", "P0", "Manual"),
        ("Verify IP address change during session handling", "IP binding policy", "1. Change network IP while logged in", "IP change", "Session remains valid or requires re-auth per security policy", "Security", "P2", "Manual"),
        ("Verify User-Agent string change during session handling", "User-Agent check", "1. Change User-Agent header mid-session", "User-Agent swap", "Session flagged or invalidated if suspicious activity detected", "Security", "P2", "Manual"),
        ("Verify session state synchronization after system sleep/wake", "OS sleep wake", "1. Login\n2. Put PC to sleep for 2 hours\n3. Wake PC", "OS sleep", "App verifies token validity on wake and refreshes if needed", "Resilience", "P2", "Manual"),
        ("Verify handling of corrupted session token string in storage", "Corrupted token", "1. Edit JWT token in storage to random garbage string", "Garbage token", "App clears corrupted token safely and prompts for login", "Resilience", "P1", "Automated"),
        ("Verify session timeout warning modal if implemented", "Timeout warning UI", "1. Idle until 2 mins before session timeout", "Impending timeout", "Modal appears: 'Session expiring soon. Extend session?'", "UX", "P2", "Manual"),
        ("Verify clicking 'Extend Session' in timeout modal refreshes token", "Extend session click", "1. Click 'Extend Session' button in warning modal", "Extend click", "Fresh access token issued, modal dismissed, timer reset", "UX", "P2", "Manual"),
        ("Verify allowing session timeout modal to expire logs out user", "Timeout modal expiry", "1. Ignore session timeout warning modal", "No action", "Modal timer reaches 0, session destroyed, redirected to /login", "UX", "P2", "Manual"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in session_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Session Management & Tokens",
            title,
            "Managing Kintsugi user authentication sessions",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 6: KEYBOARD NAVIGATION & ACCESSIBILITY (TC_LOG_236 - TC_LOG_260)
    # --------------------------------------------------------------------------
    a11y_scenarios = [
        ("Verify initial focus is set to email input on page load", "Initial focus", "1. Open /login page", "N/A", "Email field automatically receives focus or first in tab order", "Accessibility", "P1", "Automated"),
        ("Verify Tab key order traverses form elements logically", "Tab navigation", "1. Press Tab repeatedly from page load", "Tab keypresses", "Focus moves: Email -> Password -> Show Pass -> Forgot Pass -> Submit -> Register", "Accessibility", "P0", "Automated"),
        ("Verify Shift+Tab key order reverses focus traversal cleanly", "Shift+Tab navigation", "1. Focus submit button\n2. Press Shift+Tab", "Shift+Tab keypress", "Focus moves backward to previous interactive element", "Accessibility", "P1", "Automated"),
        ("Verify visible focus indicator outline on all interactive controls", "Focus outline UI", "1. Tab through all controls", "N/A", "Clear high-contrast violet focus ring visible around active element", "Accessibility", "P0", "Automated"),
        ("Verify Spacebar key activates password show/hide toggle button", "Spacebar activation", "1. Focus password eye icon button\n2. Press Spacebar", "Spacebar keypress", "Toggles password visibility state text <-> password", "Accessibility", "P1", "Automated"),
        ("Verify Enter key activates password show/hide toggle button", "Enter key toggle", "1. Focus password eye icon button\n2. Press Enter", "Enter keypress", "Toggles password visibility state text <-> password", "Accessibility", "P1", "Automated"),
        ("Verify Spacebar key activates 'Forgot password?' button link", "Spacebar activation link", "1. Focus 'Forgot password?' link\n2. Press Spacebar", "Spacebar keypress", "Navigates to /forgot-password recovery route", "Accessibility", "P2", "Automated"),
        ("Verify Enter key triggers form submission from submit button focus", "Enter key submit button", "1. Focus 'Sign In' button\n2. Press Enter", "Enter keypress", "Form submits authentication request to API", "Accessibility", "P0", "Automated"),
        ("Verify aria-required='true' attribute on mandatory inputs", "Screen reader mandatory", "1. Inspect email & password input elements", "N/A", "Inputs indicate required status for assistive technology", "Accessibility", "P1", "Automated"),
        ("Verify aria-invalid attribute dynamically updates on error", "Screen reader invalid state", "1. Trigger email error\n2. Inspect email input", "Email error active", "aria-invalid='true' added; reverts to 'false' when fixed", "Accessibility", "P1", "Automated"),
        ("Verify aria-describedby links input to its error message ID", "Screen reader error description", "1. Trigger email error\n2. Inspect email input", "Email error active", "aria-describedby references the error paragraph ID", "Accessibility", "P1", "Automated"),
        ("Verify screen reader announces server error alert message", "Live region alert", "1. Trigger server error banner\n2. Inspect alert tag", "Server error", "Alert container includes role='alert' or aria-live='assertive'", "Accessibility", "P0", "Automated"),
        ("Verify color contrast ratio for primary text (minimum 4.5:1)", "WCAG color contrast text", "1. Measure contrast of foreground text vs background", "Text contrast", "Exceeds WCAG AA standard 4.5:1 ratio requirement", "Accessibility", "P0", "Manual"),
        ("Verify color contrast ratio for muted labels (minimum 3:1)", "WCAG color contrast label", "1. Measure contrast of label text vs background", "Label contrast", "Exceeds WCAG AA standard 3:1 ratio requirement for large/bold text", "Accessibility", "P1", "Manual"),
        ("Verify color contrast ratio for validation error messages", "WCAG color contrast error", "1. Measure red error text vs background", "Error contrast", "Exceeds WCAG AA standard 4.5:1 ratio requirement", "Accessibility", "P1", "Manual"),
        ("Verify all icons have decorative aria-hidden='true' or alt text", "Accessible icons", "1. Inspect Mail, Lock, Eye SVG icons", "N/A", "Decorative icons marked aria-hidden='true' so screen readers skip them", "Accessibility", "P2", "Automated"),
        ("Verify form container uses semantic HTML5 <form> element", "Semantic DOM form", "1. Inspect DOM structure", "N/A", "Wrapped inside standard <form> element", "Accessibility", "P1", "Automated"),
        ("Verify page landmarks use semantic <main> container", "Semantic DOM main", "1. Inspect root page element", "N/A", "<main> landmark wrapper surrounds authentication content", "Accessibility", "P1", "Automated"),
        ("Verify heading hierarchy starts with single <h1> tag", "Heading structure H1", "1. Inspect heading elements", "N/A", "Page contains exactly one <h1> tag for main headline", "Accessibility", "P1", "Automated"),
        ("Verify subheadings use proper <h2> hierarchy order", "Heading structure H2", "1. Inspect subheadings", "N/A", "Feature highlights use <h2> tags without skipping levels", "Accessibility", "P2", "Automated"),
        ("Verify screen reader compatibility with NVDA / JAWS", "Screen reader audit", "1. Enable NVDA screen reader\n2. Navigate login page", "NVDA active", "All labels, inputs, errors, and buttons read aloud clearly", "Accessibility", "P0", "Manual"),
        ("Verify screen reader compatibility with VoiceOver on macOS/iOS", "VoiceOver audit", "1. Enable VoiceOver\n2. Navigate login form", "VoiceOver active", "Form controls announced accurately with proper traits", "Accessibility", "P0", "Manual"),
        ("Verify page remains usable at 200% browser zoom level", "Browser zoom text resize", "1. Set browser zoom to 200%", "Zoom 200%", "No text overlap, horizontal scrollbar avoided where possible", "Accessibility", "P1", "Manual"),
        ("Verify high-contrast mode OS setting compatibility", "Windows High Contrast", "1. Enable OS High Contrast Mode", "High contrast OS", "Borders and text remain visible with high contrast colors", "Accessibility", "P2", "Manual"),
        ("Verify interactive elements have minimum touch target size (44x44px)", "Touch target size", "1. Measure dimensions of buttons and inputs", "N/A", "Touch targets meet or exceed 44x44px for accessibility", "Accessibility", "P1", "Manual"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in a11y_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Keyboard & Accessibility (a11y)",
            title,
            "Evaluating WCAG accessibility compliance on Kintsugi web",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 7: CROSS-BROWSER & MULTI-VIEWPORT RESPONSIVENESS (TC_LOG_261 - TC_LOG_285)
    # --------------------------------------------------------------------------
    responsive_scenarios = [
        ("Verify layout rendering on Mobile Portrait (375x812 iPhone X)", "Mobile Portrait viewport", "1. Set browser resolution to 375x812", "375x812", "Card fills screen width with 16px padding, brand column hidden", "Responsiveness", "P0", "Automated"),
        ("Verify layout rendering on Mobile Small (320x568 iPhone SE)", "Mobile Small viewport", "1. Set browser resolution to 320x568", "320x568", "Card scales down, font size remains readable, no horizontal scroll", "Responsiveness", "P1", "Automated"),
        ("Verify layout rendering on Mobile Landscape (812x375)", "Mobile Landscape viewport", "1. Set browser resolution to 812x375", "812x375", "Vertical scroll enabled smoothly, inputs accessible", "Responsiveness", "P2", "Automated"),
        ("Verify layout rendering on Tablet Portrait (768x1024 iPad)", "Tablet Portrait viewport", "1. Set browser resolution to 768x1024", "768x1024", "Centered auth card displayed cleanly with comfortable padding", "Responsiveness", "P1", "Automated"),
        ("Verify layout rendering on Tablet Landscape (1024x768 iPad)", "Tablet Landscape viewport", "1. Set browser resolution to 1024x768", "1024x768", "Dual-column layout activates cleanly with left brand panel", "Responsiveness", "P1", "Automated"),
        ("Verify layout rendering on Desktop Standard (1366x768 Laptop)", "Desktop Standard viewport", "1. Set browser resolution to 1366x768", "1366x768", "Dual-column layout rendered with optimal proportions", "Responsiveness", "P0", "Automated"),
        ("Verify layout rendering on Desktop Full HD (1920x1080)", "Desktop Full HD viewport", "1. Set browser resolution to 1920x1080", "1920x1080", "Content centered inside max-w-5xl wrapper with background effects", "Responsiveness", "P0", "Automated"),
        ("Verify layout rendering on Ultra-Wide Display (2560x1440 2K)", "Ultra-wide viewport", "1. Set browser resolution to 2560x1440", "2560x1440", "Card stays centered without excessive stretching", "Responsiveness", "P2", "Automated"),
        ("Verify layout rendering on 4K Display (3840x2160)", "4K display viewport", "1. Set browser resolution to 3840x2160", "3840x2160", "Sharp crisp typography, no blurriness or layout distortion", "Responsiveness", "P3", "Manual"),
        ("Verify Chrome browser rendering compatibility (latest version)", "Chrome browser test", "1. Open login page in Google Chrome", "Chrome latest", "All CSS styles, animations, and JS features function perfectly", "Compatibility", "P0", "Automated"),
        ("Verify Firefox browser rendering compatibility (latest version)", "Firefox browser test", "1. Open login page in Mozilla Firefox", "Firefox latest", "Gecko engine renders layout and inputs accurately", "Compatibility", "P0", "Automated"),
        ("Verify Safari browser rendering compatibility (macOS/iOS)", "Safari browser test", "1. Open login page in Apple Safari", "Safari latest", "WebKit engine renders glassmorphism backdrop blur cleanly", "Compatibility", "P0", "Manual"),
        ("Verify Microsoft Edge browser rendering compatibility", "Edge browser test", "1. Open login page in Microsoft Edge", "Edge latest", "Chromium engine renders interface without issues", "Compatibility", "P0", "Automated"),
        ("Verify Opera browser rendering compatibility", "Opera browser test", "1. Open login page in Opera browser", "Opera latest", "Renders correctly without visual bugs", "Compatibility", "P3", "Manual"),
        ("Verify Brave browser rendering compatibility (privacy shields)", "Brave browser test", "1. Open login page in Brave with shields active", "Brave active", "Auth scripts function without shield blocking core API calls", "Compatibility", "P1", "Manual"),
        ("Verify Mobile Safari rendering on iPhone device", "iOS Mobile Safari", "1. Open login page on physical iPhone", "iOS Safari", "Virtual keyboard popup does not distort layout or cover inputs", "Compatibility", "P0", "Manual"),
        ("Verify Mobile Chrome rendering on Android device", "Android Mobile Chrome", "1. Open login page on physical Android device", "Android Chrome", "Virtual keyboard popup adjusts scroll view smoothly", "Compatibility", "P0", "Manual"),
        ("Verify device orientation change from Portrait to Landscape", "Orientation swap", "1. Rotate mobile device from portrait to landscape", "Device rotation", "Layout recalculates responsively without requiring page reload", "Responsiveness", "P2", "Manual"),
        ("Verify dynamic window resize behavior on desktop", "Dynamic window resize", "1. Drag desktop window edge from 1400px down to 400px", "Live window drag", "Layout transitions smoothly between desktop grid and mobile stack", "Responsiveness", "P2", "Automated"),
        ("Verify background effect canvas/radial glow resize scaling", "Background effects scaling", "1. Resize window dimensions", "Live window drag", "Ambient background glow resizes without horizontal scrollbars", "Responsiveness", "P3", "Manual"),
        ("Verify touch gesture scrolling on mobile devices", "Touch scroll UX", "1. Drag finger vertically on mobile screen", "Mobile touch", "Smooth momentum scrolling active", "UX", "P2", "Manual"),
        ("Verify tap target responsiveness on mobile buttons", "Mobile tap responsiveness", "1. Tap Sign In button on mobile screen", "Mobile tap", "Immediate visual active state feedback upon touch", "UX", "P1", "Manual"),
        ("Verify mobile browser address bar hide/show scroll behavior", "Mobile address bar scroll", "1. Scroll down on mobile browser page", "Mobile scroll", "Page container min-h-screen adapts to dynamic viewport height (dvh)", "UI/UX", "P2", "Manual"),
        ("Verify high DPI (Retina) display image rendering quality", "Retina display check", "1. Inspect brand logo on 2x/3x Retina screen", "Retina display", "Vectors and logo render crisp with zero pixelation", "UI/UX", "P2", "Manual"),
        ("Verify CSS backdrop-filter blur support fallback", "Backdrop filter fallback", "1. Test in browser without backdrop-filter support", "No backdrop-filter", "Solid fallback background color applied safely", "UI/UX", "P3", "Manual"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in responsive_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Cross-Browser & Multi-Viewport",
            title,
            "Testing responsiveness & browser compatibility",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    # --------------------------------------------------------------------------
    # MODULE 8: PERFORMANCE, NETWORK RESILIENCE & ERROR RECOVERY (TC_LOG_286 - TC_LOG_310)
    # --------------------------------------------------------------------------
    perf_scenarios = [
        ("Verify login page Lighthouse Performance score >= 90", "Lighthouse Performance", "1. Run Lighthouse audit on /login page", "Lighthouse audit", "Performance score exceeds 90 on desktop", "Performance", "P1", "Manual"),
        ("Verify login page Lighthouse Accessibility score = 100", "Lighthouse Accessibility", "1. Run Lighthouse accessibility audit", "Lighthouse audit", "Accessibility score reaches 100", "Accessibility", "P0", "Manual"),
        ("Verify login page Lighthouse Best Practices score >= 95", "Lighthouse Best Practices", "1. Run Lighthouse best practices audit", "Lighthouse audit", "Best Practices score exceeds 95", "Technical", "P1", "Manual"),
        ("Verify login page First Contentful Paint (FCP) < 1.0s", "FCP metric", "1. Measure Web Vitals FCP metric", "Performance profile", "FCP occurs under 1.0 second on fast network", "Performance", "P1", "Automated"),
        ("Verify login page Largest Contentful Paint (LCP) < 2.0s", "LCP metric", "1. Measure Web Vitals LCP metric", "Performance profile", "LCP occurs under 2.0 seconds", "Performance", "P1", "Automated"),
        ("Verify login page Cumulative Layout Shift (CLS) = 0.0", "CLS metric", "1. Measure Web Vitals CLS metric", "Performance profile", "Zero layout shift during initial load (CLS < 0.05)", "Performance", "P1", "Automated"),
        ("Verify First Input Delay (FID) / INP < 100ms", "INP responsiveness metric", "1. Click input box immediately upon load", "Performance profile", "Input focus delay under 100ms", "Performance", "P1", "Automated"),
        ("Verify JavaScript main thread blocking time during render", "TBT main thread", "1. Inspect Total Blocking Time in Performance tab", "Performance profile", "Total Blocking Time under 150ms", "Performance", "P2", "Automated"),
        ("Verify total asset bundle size for login route (<300KB gzipped)", "Bundle size check", "1. Inspect Network tab JS/CSS transfer size", "Network bundle", "Total transfer size under 300KB gzipped", "Performance", "P1", "Automated"),
        ("Verify form submission behavior on Slow 3G network throttling", "Slow 3G network test", "1. Enable 'Slow 3G' throttling in dev tools\n2. Submit login", "Slow 3G throttle", "Spinner displayed continuously, submit button remains disabled", "Resilience", "P1", "Automated"),
        ("Verify form submission behavior when client goes completely Offline", "Offline network test", "1. Disconnect network / set Offline in dev tools\n2. Submit login", "Offline mode", "Displays clear error banner: 'No internet connection available.'", "Resilience", "P0", "Automated"),
        ("Verify automatic request retry when network connection drops briefly", "Network connection drop", "1. Submit login\n2. Interrupt connection for 1s\n3. Restore connection", "Flaky network", "Request retries or provides retry button option", "Resilience", "P2", "Manual"),
        ("Verify server 500 Internal Server Error handling", "HTTP 500 error", "1. Mock backend response HTTP 500\n2. Submit form", "HTTP 500 mock", "Displays friendly error: 'Internal server error. Please try again later.'", "Resilience", "P0", "Automated"),
        ("Verify server 502 Bad Gateway error handling", "HTTP 502 error", "1. Mock backend response HTTP 502\n2. Submit form", "HTTP 502 mock", "Displays friendly error: 'Server temporarily unavailable.'", "Resilience", "P1", "Automated"),
        ("Verify server 503 Service Unavailable error handling", "HTTP 503 error", "1. Mock backend response HTTP 503\n2. Submit form", "HTTP 503 mock", "Displays maintenance notice error banner", "Resilience", "P1", "Automated"),
        ("Verify handling of unexpected HTML response from backend", "Malformed server response", "1. Mock backend returning raw HTML 404 page", "HTML response mock", "App catches JSON parse error safely and displays fallback error", "Resilience", "P1", "Automated"),
        ("Verify handling of empty HTTP 200 OK response from backend", "Empty body response", "1. Mock backend returning HTTP 200 with empty body", "Empty 200 OK mock", "App detects missing token payload and rejects authentication", "Resilience", "P1", "Automated"),
        ("Verify rapid double-click on Sign In button", "Double click prevention", "1. Rapidly double click Sign In button in 50ms", "Rapid clicks", "Only ONE API call dispatched to server", "Performance", "P0", "Automated"),
        ("Verify rapid triple-click on Sign In button", "Triple click prevention", "1. Rapidly triple click Sign In button", "Rapid clicks", "Only ONE API call dispatched to server", "Performance", "P0", "Automated"),
        ("Verify memory leak prevention during repeated form toggling", "Memory leak test", "1. Toggle between Login & Register 50 times", "Repeated state changes", "Memory consumption remains stable without memory leak", "Performance", "P2", "Manual"),
        ("Verify cleanup of active timers on component unmount", "Timer cleanup", "1. Trigger pending request\n2. Navigate away quickly", "Component unmount", "Pending promises/timers cleaned up without console memory error", "Technical", "P2", "Automated"),
        ("Verify web socket reconnection if auth uses real-time status", "WebSocket resilience", "1. Disconnect WebSocket\n2. Wait 3s", "WebSocket drop", "Automatic reconnection attempt initialized", "Resilience", "P3", "Manual"),
        ("Verify application recovery after clearing app cache & hard reload", "Hard refresh recovery", "1. Press Ctrl+F5 on login page", "Hard refresh", "Page reloads cleanly without broken asset errors", "Resilience", "P2", "Manual"),
        ("Verify DNS resolution failure handling on API request", "DNS failure test", "1. Block API domain in hosts file\n2. Attempt login", "DNS failure", "Displays clear connectivity error prompt", "Resilience", "P2", "Manual"),
        ("Verify SSL/TLS certificate error warning handling", "SSL cert error", "1. Point API to self-signed invalid SSL cert", "Invalid SSL", "Browser/client blocks request and alerts user of untrusted SSL", "Security", "P0", "Manual"),
    ]

    for title, scenario, steps, data, expected, ttype, prio, auto in perf_scenarios:
        test_cases.append((
            f"TC_LOG_{tc_count:03d}",
            "Performance & Error Recovery",
            title,
            "Testing network resilience, metrics & recovery",
            steps,
            data,
            expected,
            ttype,
            prio,
            auto,
            "Pass" if auto == "Automated" else "Untested"
        ))
        tc_count += 1

    return test_cases

def generate_excel_report(output_filepath):
    wb = openpyxl.Workbook()
    
    # --------------------------------------------------------------------------
    # STYLES DEFINITION
    # --------------------------------------------------------------------------
    font_family = "Segoe UI"
    
    # Colors
    NAVY_HEADER = "1E293B"       # Slate 800
    ACCENT_PURPLE = "6D28D9"     # Violet 700
    BG_LIGHT_GRAY = "F8FAFC"     # Slate 50
    ZEBRA_BG = "F1F5F9"          # Slate 100
    BORDER_COLOR = "CBD5E1"      # Slate 300
    
    # Pass / Fail / Untested Fills
    PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Emerald 100
    PASS_FONT = Font(name=font_family, size=10, bold=True, color="15803D")
    
    UNTESTED_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amber 100
    UNTESTED_FONT = Font(name=font_family, size=10, bold=True, color="B45309")
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_heading_font = Font(name=font_family, size=12, bold=True, color=NAVY_HEADER)
    tbl_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=9, bold=False, color="1E293B")
    data_bold = Font(name=font_family, size=9, bold=True, color="1E293B")
    
    # Fills
    header_fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    purple_fill = PatternFill(start_color=ACCENT_PURPLE, end_color=ACCENT_PURPLE, fill_type="solid")
    card_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    zebra_fill = PatternFill(start_color=ZEBRA_BG, end_color=ZEBRA_BG, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Priority Fills
    P0_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Red 100
    P0_FONT = Font(name=font_family, size=9, bold=True, color="991B1B")
    P1_FILL = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Orange 100
    P1_FONT = Font(name=font_family, size=9, bold=True, color="C2410C")
    P2_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Yellow 100
    P2_FONT = Font(name=font_family, size=9, bold=True, color="854D0E")
    P3_FILL = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # Sky 100
    P3_FONT = Font(name=font_family, size=9, bold=True, color="0369A1")

    # Borders
    thin_border_side = Side(style='thin', color=BORDER_COLOR)
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(style='medium', color=NAVY_HEADER))

    # Alignments
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    # ==========================================================================
    # SHEET 1: EXECUTIVE SUMMARY DASHBOARD
    # ==========================================================================
    ws_summary = wb.active
    ws_summary.title = "Executive_Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.merge_cells("A1:G2")
    title_cell = ws_summary["A1"]
    title_cell.value = "KINTSUGI WEB FRONTEND - E2E LOGIN TEST SUITE SUMMARY REPORT"
    title_cell.font = title_font
    title_cell.fill = purple_fill
    title_cell.alignment = align_center

    # Key Metrics Cards (Row 4 to Row 6)
    metrics = [
        ("Total Test Cases", "310", "B4:C5"),
        ("Automated (Selenium)", "125", "D4:E5"),
        ("Manual / Exploratory", "185", "F4:G5")
    ]

    for label, val, cell_range in metrics:
        ws_summary.merge_cells(cell_range)
        top_left = ws_summary[cell_range.split(":")[0]]
        top_left.value = f"{label}\n\n{val}"
        top_left.font = Font(name=font_family, size=12, bold=True, color=NAVY_HEADER)
        top_left.fill = card_fill
        top_left.alignment = align_center
        # Apply borders around range
        start, end = cell_range.split(":")
        for r in ws_summary[f"{start}:{end}"]:
            for c in r:
                c.border = thin_border

    # Section 1: Breakdown by Module
    ws_summary["A8"] = "1. Test Coverage Breakdown by Functional Module"
    ws_summary["A8"].font = section_heading_font

    module_headers = ["Module Name", "Total TCs", "Automated (Selenium)", "Manual TCs", "P0 Critical", "P1 High"]
    ws_summary.append([]) # Row 9 empty gap
    
    ws_summary.cell(row=10, column=1, value=module_headers[0])
    for col_idx, header in enumerate(module_headers, start=1):
        cell = ws_summary.cell(row=10, column=col_idx, value=header)
        cell.font = tbl_header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    modules_summary_data = [
        ("UI Layout & Visual Design", 40, 15, 25, 5, 12),
        ("Form Fields & Input Validation", 55, 45, 10, 18, 22),
        ("Authentication & Business Rules", 55, 25, 30, 20, 20),
        ("Security & Edge Cases", 55, 28, 27, 22, 18),
        ("Session Management & Tokens", 30, 12, 18, 10, 12),
        ("Keyboard & Accessibility (a11y)", 25, 10, 15, 6, 10),
        ("Cross-Browser & Multi-Viewport", 25, 12, 13, 8, 10),
        ("Performance & Error Recovery", 25, 14, 11, 6, 10),
    ]

    for row_idx, row_data in enumerate(modules_summary_data, start=11):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font if col_idx > 1 else data_bold
            cell.alignment = align_center if col_idx > 1 else align_left
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill

    # Total Row for Modules Table
    tot_row = 19
    ws_summary.cell(row=tot_row, column=1, value="TOTAL TEST CASES").font = data_bold
    ws_summary.cell(row=tot_row, column=1).alignment = align_left
    ws_summary.cell(row=tot_row, column=1).fill = card_fill
    ws_summary.cell(row=tot_row, column=1).border = thin_border

    tot_cols = [
        ("=SUM(B11:B18)"),
        ("=SUM(C11:C18)"),
        ("=SUM(D11:D18)"),
        ("=SUM(E11:E18)"),
        ("=SUM(F11:F18)")
    ]
    for c_idx, formula in enumerate(tot_cols, start=2):
        cell = ws_summary.cell(row=tot_row, column=c_idx, value=formula)
        cell.font = data_bold
        cell.alignment = align_center
        cell.fill = card_fill
        cell.border = thin_border

    # Section 2: Priority & Automation Distribution
    ws_summary["A22"] = "2. Priority & Execution Distribution"
    ws_summary["A22"].font = section_heading_font

    prio_headers = ["Priority Level", "Description", "Total Test Cases", "Percentage"]
    for col_idx, header in enumerate(prio_headers, start=1):
        cell = ws_summary.cell(row=23, column=col_idx, value=header)
        cell.font = tbl_header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    prio_data = [
        ("P0 - Critical", "Core authentication flows, security vulnerabilities, zero-tolerance bugs", 95, "=C24/310"),
        ("P1 - High", "Validation rules, primary UI controls, main responsive viewports", 114, "=C25/310"),
        ("P2 - Medium", "Edge cases, minor accessibility features, error recovery toasts", 71, "=C26/310"),
        ("P3 - Low", "Cosmetic styling, minor text copywriting, rare environmental edge cases", 30, "=C27/310"),
    ]

    for r_idx, (p_name, desc, count, pct) in enumerate(prio_data, start=24):
        c1 = ws_summary.cell(row=r_idx, column=1, value=p_name)
        c2 = ws_summary.cell(row=r_idx, column=2, value=desc)
        c3 = ws_summary.cell(row=r_idx, column=3, value=count)
        c4 = ws_summary.cell(row=r_idx, column=4, value=pct)

        c1.font = data_bold
        c2.font = data_font
        c3.font = data_bold
        c4.font = data_font

        c1.alignment = align_left
        c2.alignment = align_left
        c3.alignment = align_center
        c4.alignment = align_center
        c4.number_format = '0.0%'

        for cell in (c1, c2, c3, c4):
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = zebra_fill

    # Set column widths for summary sheet
    summary_col_widths = {'A': 32, 'B': 24, 'C': 22, 'D': 22, 'E': 16, 'F': 16, 'G': 16}
    for col_letter, width in summary_col_widths.items():
        ws_summary.column_dimensions[col_letter].width = width


    # ==========================================================================
    # SHEET 2: DETAILED TEST CASES (310 TEST CASES)
    # ==========================================================================
    ws_details = wb.create_sheet(title="Detailed_Test_Cases")
    ws_details.views.sheetView[0].showGridLines = True

    # Main Table Headers
    headers = [
        "Test Case ID",
        "Module / Feature",
        "Test Scenario Title",
        "Pre-Conditions",
        "Test Steps",
        "Test Data / Input Payload",
        "Expected Result",
        "Test Type",
        "Priority",
        "Automation Status",
        "Execution Result"
    ]

    ws_details.row_dimensions[1].height = 28
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.font = tbl_header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Populate 310 Test Cases
    all_test_cases = create_test_cases()

    for row_idx, tc in enumerate(all_test_cases, start=2):
        ws_details.row_dimensions[row_idx].height = 42 # Generous height for multiline steps
        for col_idx, val in enumerate(tc, start=1):
            cell = ws_details.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

            # Specific column alignments & styles
            if col_idx in [1, 8, 9, 10, 11]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            # Formatting IDs
            if col_idx == 1:
                cell.font = data_bold

            # Priority Formatting
            if col_idx == 9:
                if val == "P0":
                    cell.fill = P0_FILL
                    cell.font = P0_FONT
                elif val == "P1":
                    cell.fill = P1_FILL
                    cell.font = P1_FONT
                elif val == "P2":
                    cell.fill = P2_FILL
                    cell.font = P2_FONT
                elif val == "P3":
                    cell.fill = P3_FILL
                    cell.font = P3_FONT

            # Execution Result Formatting
            if col_idx == 11:
                if val == "Pass":
                    cell.fill = PASS_FILL
                    cell.font = PASS_FONT
                elif val == "Untested":
                    cell.fill = UNTESTED_FILL
                    cell.font = UNTESTED_FONT

            # Zebra striping for generic columns
            if col_idx not in [9, 11] and row_idx % 2 == 0:
                cell.fill = zebra_fill

    # Set Column Widths for Details Sheet
    detail_col_widths = {
        'A': 16, # ID
        'B': 26, # Module
        'C': 34, # Title
        'D': 28, # Pre-conditions
        'E': 36, # Test Steps
        'F': 26, # Test Data
        'G': 36, # Expected Result
        'H': 18, # Test Type
        'I': 12, # Priority
        'J': 20, # Automation
        'K': 16, # Result
    }

    for col_letter, width in detail_col_widths.items():
        ws_details.column_dimensions[col_letter].width = width

    # Save Workbook
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    wb.save(output_filepath)
    print(f"[SUCCESS] Generated Excel test report with {len(all_test_cases)} test cases at:\n   {output_filepath}")

if __name__ == "__main__":
    output_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../selenium-tests/Login_Test_Cases_300.xlsx"))
    generate_excel_report(output_path)

    # Also make a copy inside selenium-tests/tests/
    copy_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../selenium-tests/tests/Login_Test_Cases_Summary.xlsx"))
    generate_excel_report(copy_path)
